#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_import.py — 倉庫庫存表 → MySQL 匯入 SQL 產生器
"""
import openpyxl
import re
import json
import os
from datetime import datetime, date

# --- 設定 ---
EXCEL_FILE = r'Z:/BOM/ERP/AI_倉庫庫存表-雲端.xlsm'
OUTPUT_SQL  = 'import_transactions.sql'
OUTPUT_JSON = 'import_report.json'

# 工作表名稱 → item_type（對應 stock_item_categories.category_id）
SHEET_ITEM_TYPE = {
    '半成品': 2,   # 半成品
}
DEFAULT_ITEM_TYPE = 1  # 其他工作表視為一般件/成品

OP_MAP = {
    '入': 'in', '進': 'in', '收': 'in', '回': 'in',
    '出': 'out', '領': 'out', '發': 'out', '扣': 'out',
    '移': 'move', '換': 'move', '儲位': 'move', '儲位變更': 'move',
    '盤': 'count', '清': 'count',
    '調整': 'adjust', '遺': 'adjust', '失': 'adjust', '廢': 'adjust'
}

# ── 各工作表欄位標題 → DB欄位對應 ────────────────────────────────────────
# 半成品工作表格式（無製造日期、無包裝箱、無售價、無保存年限欄）：
#   A=客戶 B=件號 C=一廠位置 D=總數 E=出入庫 F=備註1 G=備註2
#   H=產品品名 I=產品規格 J=製令單號  O欄=無售價
#
# 其他工作表標準格式：
#   A=客戶 B=件號 C=倉區儲位 D=總數 E=出入庫 F=製造日期 G=庫存日期
#   H=保存年限 I=備註1 J=備註2 K=產品品名 L=產品規格 M=製令單號 N=包裝箱 O=售價
#
# 標題關鍵字對應（strip後比對）
COL_MAP = {
    '客戶':       'client_name_col',
    '件號':       'part_col',
    '總數':       'qty_col',
    '出入庫':     'txn_col',       # comment欄（E欄，col=5）
    '製造日期':   'mfg_date_col',
    '備註1':      'remark1_col',
    '備註2':      'remark2_col',
    '包裝箱':     'package_box_col',
    '製令單號':   'bom_ref_col',
    '售價':       'unit_price_col',
    '倉區儲位':   'location_col',
    '一廠位置\n(排-格-箱)': 'location_col',
    '一廠位置':   'location_col',
}

def detect_columns(ws):
    """讀取第3列標題，動態建立欄號對應 dict"""
    cols = {}
    for c in range(1, 20):
        raw = ws.cell(row=3, column=c).value
        if raw is None:
            continue
        key = str(raw).strip().split('\n')[0].strip()  # 取第一行、去空白
        for pattern, field in COL_MAP.items():
            # 支援多行標題（備註2(同料號位置) 等）
            if key.startswith(pattern.split('\n')[0].strip()):
                cols[field] = c
                break
    # 售價欄若標題已被刪除，固定補O欄（col=15），半成品工作表不補
    return cols

def parse_date_value(val):
    """解析 Excel 日期欄位：可能是整數(20230301)或 date 物件"""
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    try:
        s = str(int(val))
        if len(s) == 8:
            return date(int(s[0:4]), int(s[4:6]), int(s[6:8]))
    except:
        pass
    return None

def parse_line(line, default_loc=""):
    raw_text = line.strip()
    clean_line = raw_text.replace(' ', '')
    if not clean_line:
        return None

    date_match = re.search(r'(\d{4})/(\d{1,3})/(\d{1,3})', clean_line)
    if not date_match:
        return {'error': 'Invalid Date Format', 'raw': raw_text}
    y, m, d_str = date_match.groups()
    try:
        parsed_date = date(int(y), int(m[:2]), int(d_str[:2]))
    except:
        return {'error': 'Date Value Error', 'raw': raw_text}

    qtys = re.findall(r'\((\d+\.?\d*)\)', clean_line)
    op_type, op_zh = 'adjust', 'Change'
    for key, val in OP_MAP.items():
        if key in clean_line:
            op_type = val
            op_zh = '移' if val == 'move' else key
            break

    loc_from, loc_to = default_loc, default_loc
    if op_type == 'move':
        move_match = re.search(r'\((\d+\.?\d*)\)([^→>\s(]+)[→> -]+([^→>\s(]+)', clean_line)
        if move_match:
            loc_from, loc_to = move_match.group(2).strip(), move_match.group(3).strip()
        else:
            alt_match = re.search(r'(?:至|到|變更為)([^(\n]+)', raw_text)
            if alt_match:
                loc_to = alt_match.group(1).strip()

    qty, stock_after = 0.0, 0.0
    if len(qtys) >= 2:
        qty, stock_after = float(qtys[0]), float(qtys[-1])
    elif len(qtys) == 1:
        if '庫存' in clean_line or '餘' in clean_line:
            stock_after = float(qtys[0])
        else:
            qty = float(qtys[0])

    return {
        'date': parsed_date.isoformat(), 'op': op_type, 'op_zh': op_zh,
        'qty': qty, 'stock_after': stock_after, 'loc_from': loc_from, 'loc_to': loc_to,
        'display_loc': f"{loc_from} -> {loc_to}" if loc_from != loc_to else loc_to,
        'raw': raw_text, 'remark': raw_text.split(')')[-1] if ')' in raw_text else ''
    }

def sql_str(val):
    """對字串做 SQL 跳脫並包上引號；None → NULL"""
    if val is None or val == '':
        return 'NULL'
    return "'" + str(val).replace("'", "''") + "'"

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: File not found {EXCEL_FILE}")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, keep_vba=False, data_only=True)
    except Exception as e:
        print(f"Error: Load failed {e}")
        return

    report = {
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'summary': {'total_items': 0, 'total_transactions': 0, 'parse_errors': 0,
                    'quantity_mismatches': 0, 'sequence_warnings': 0},
        'errors': [], 'mismatches': [], 'warnings': [], 'items': []
    }
    sql_lines = ["START TRANSACTION;\n"]
    total_ok, total_error = 0, 0

    for ws in wb.worksheets:
        ws_name = ws.title
        item_type = SHEET_ITEM_TYPE.get(ws_name, DEFAULT_ITEM_TYPE)
        is_hanpin = (ws_name == '半成品')

        # ── 動態偵測欄位對應 ──────────────────────────────────────────────
        cols = detect_columns(ws)
        # 出入庫欄固定 E=5（comment欄）
        txn_col = 5
        # 售價：半成品無售價；其他工作表固定 O=15
        unit_price_col = None if is_hanpin else cols.get('unit_price_col', 15)

        for row_num in range(4, ws.max_row + 1):
            # 半成品：E欄有comment才處理；其他工作表同樣判斷 E 欄 comment
            cell_e = ws.cell(row=row_num, column=txn_col)
            if not cell_e.comment:
                continue

            # ── 讀取各欄資料 ────────────────────────────────────────────
            def cv(field, default=''):
                c = cols.get(field)
                if c is None:
                    return default
                v = ws.cell(row=row_num, column=c).value
                return v

            client_name = str(cv('client_name_col') or '').strip()
            d_id_raw    = ws.cell(row=row_num, column=2).value   # B欄固定=件號
            d_id        = str(d_id_raw or '').strip()

            if not d_id or d_id in ('件號', '倉庫庫存表', 'None'):
                continue

            # D欄固定=總數
            raw_qty = ws.cell(row=row_num, column=4).value
            try:
                ex_qty = float(raw_qty) if raw_qty is not None else 0.0
            except (ValueError, TypeError):
                num_str = re.sub(r'[^\d.-]', '', str(raw_qty))
                ex_qty = float(num_str) if num_str else 0.0

            # C欄固定=儲位
            default_loc = str(ws.cell(row=row_num, column=3).value or '').strip()

            # 製造日期（標準格式工作表 F 欄；半成品無此欄）
            mfg_date = None
            if not is_hanpin and cols.get('mfg_date_col'):
                mfg_date = parse_date_value(ws.cell(row=row_num, column=cols['mfg_date_col']).value)

            # 備註1 / 備註2
            remark1 = str(cv('remark1_col') or '').strip() or None
            # 備註2 可能含換行（同料號位置說明），取第一行
            remark2_raw = cv('remark2_col')
            remark2 = str(remark2_raw or '').strip().split('\n')[0].strip() or None

            # 包裝箱
            package_box = str(cv('package_box_col') or '').strip() or None

            # 製令單號
            bom_ref_val = str(cv('bom_ref_col') or '').strip() or None
            # 製令單號可能含換行（多筆），取第一行
            if bom_ref_val:
                bom_ref_val = bom_ref_val.split('\n')[0].strip() or None

            # 售價
            unit_price_val = None
            if unit_price_col:
                up_raw = ws.cell(row=row_num, column=unit_price_col).value
                if up_raw is not None:
                    try:
                        unit_price_val = float(up_raw)
                    except (ValueError, TypeError):
                        unit_price_val = None

            item_rec = {
                'row': f"{ws_name}-L{row_num}",
                'client_name': client_name,
                'part': d_id,
                'd_id': d_id,
                'location': default_loc,
                'excel_qty': ex_qty,
                'item_type': item_type,
                'transactions': [], 'errors': [], 'mismatch': None
            }

            raw_lines = [l.strip() for l in cell_e.comment.text.split('\n') if l.strip()]
            if not raw_lines:
                continue

            has_header = (
                not re.search(r'\d{4}/\d{1,3}/\d{1,3}', raw_lines[0])
                and '庫存' in raw_lines[0]
            )

            parsed_res, first_found = [], False
            for idx, ln in enumerate(raw_lines):
                if idx == 0 and has_header:
                    continue
                res = parse_line(ln, default_loc)
                if not res:
                    continue
                if 'error' in res:
                    err = {'row': item_rec['row'], 'part': item_rec['part'], **res}
                    report['errors'].append(err)
                    item_rec['errors'].append(err)
                    total_error += 1
                else:
                    if not first_found and has_header:
                        res['op'], res['op_zh'] = 'in', '入'
                    res['location'] = res['display_loc']
                    parsed_res.append(res)
                    item_rec['transactions'].append(res)
                    first_found = True

            if ex_qty == 0.0 and parsed_res:
                ex_qty = parsed_res[-1]['stock_after']
                item_rec['excel_qty'] = ex_qty

            final_qty = int(ex_qty)

            # 入庫日期：取第一筆 'in' 交易日期
            stock_date = None
            for p in parsed_res:
                if p['op'] == 'in':
                    stock_date = p['date']
                    break

            # SQL 跳脫
            d_id_s       = d_id.replace("'", "''")
            loc_s        = default_loc.replace("'", "''")
            client_s     = client_name.replace("'", "''")
            stock_date_s = f"'{stock_date}'" if stock_date else 'NULL'
            mfg_date_s   = f"'{mfg_date.isoformat()}'" if mfg_date else 'NULL'
            remark1_s    = sql_str(remark1)
            remark2_s    = sql_str(remark2)
            pkg_box_s    = sql_str(package_box)
            bom_ref_s    = sql_str(bom_ref_val)
            unit_price_s = str(unit_price_val) if unit_price_val is not None else 'NULL'

            # ── 0. 確保 customer_list 存在（若無則自動新增） ────────────────
            if client_name:
                sql_lines.append(
                    f"INSERT INTO customer_list (customer_id, customer, Created_By) "
                    f"SELECT "
                    f"  LPAD(FLOOR(RAND()*900000000000 + 100000000000), 11, '0'), "
                    f"  '{client_s}', "
                    f"  'IMPORT' "
                    f"FROM DUAL "
                    f"WHERE NOT EXISTS "
                    f"  (SELECT 1 FROM customer_list WHERE TRIM(customer) = '{client_s}');"
                )

            # ── 1. 確保 d_setting 存在（若無則自動新增） ────────────────────
            sql_lines.append(
                f"INSERT INTO d_setting (D_Setting_Id, Customer_Id, Created_By) "
                f"SELECT "
                f"  '{d_id_s}', "
                f"  (SELECT customer_id FROM customer_list WHERE TRIM(customer) = '{client_s}' LIMIT 1), "
                f"  'IMPORT' "
                f"FROM DUAL "
                f"WHERE NOT EXISTS "
                f"  (SELECT 1 FROM d_setting WHERE D_Setting_Id = '{d_id_s}');"
            )

            # ── 2. 確保儲位存在 ──────────────────────────────────────────────
            locations_to_ensure = set()
            if default_loc.strip():
                locations_to_ensure.add(default_loc.strip())
            for p in parsed_res:
                for lk in ('loc_from', 'loc_to'):
                    v = p.get(lk, '').strip()
                    if v:
                        locations_to_ensure.add(v)

            for loc in locations_to_ensure:
                ls = loc.replace("'", "''")
                sql_lines.append(
                    f"INSERT INTO stock_locations (location_code, location_name) "
                    f"SELECT '{ls}', '{ls}' FROM DUAL "
                    f"WHERE NOT EXISTS "
                    f"  (SELECT 1 FROM stock_locations WHERE location_code = '{ls}');"
                )

            # ── 3. INSERT stock_items（若不存在） ────────────────────────────
            sql_lines.append(
                f"INSERT INTO stock_items "
                f"(d_id, d_setting_id, client_name, client_id, item_type, "
                f" storage_location, location_id, qty, stock_date, mfg_date, "
                f" expire_years, remark1, remark2, package_box, bom_ref, unit_price) "
                f"SELECT "
                f"  '{d_id_s}', "
                f"  (SELECT d_id FROM d_setting WHERE D_Setting_Id = '{d_id_s}' LIMIT 1), "
                f"  '{client_s}', "
                f"  (SELECT customer_id FROM customer_list WHERE TRIM(customer) = '{client_s}' LIMIT 1), "
                f"  {item_type}, "
                f"  {sql_str(default_loc if default_loc else None)}, "
                f"  (SELECT location_id FROM stock_locations WHERE location_code = '{loc_s}' LIMIT 1), "
                f"  {final_qty}, "
                f"  {stock_date_s}, "
                f"  {mfg_date_s}, "
                f"  10, "
                f"  {remark1_s}, "
                f"  {remark2_s}, "
                f"  {pkg_box_s}, "
                f"  {bom_ref_s}, "
                f"  {unit_price_s} "
                f"FROM DUAL "
                f"WHERE NOT EXISTS (SELECT 1 FROM stock_items WHERE d_id = '{d_id_s}');"
            )

            # ── 4. UPDATE stock_items（補齊所有欄位 + 更新 qty） ─────────────
            set_parts = [
                f"d_setting_id = IFNULL(d_setting_id, "
                f"  (SELECT d_id FROM d_setting WHERE D_Setting_Id = '{d_id_s}' LIMIT 1))",
                f"client_id = IFNULL(client_id, "
                f"  (SELECT customer_id FROM customer_list WHERE TRIM(customer) = '{client_s}' LIMIT 1))",
                f"client_name = '{client_s}'",
                f"item_type = {item_type}",
                f"qty = {final_qty}",
                f"expire_years = IFNULL(expire_years, 10)",
            ]
            if default_loc:
                set_parts += [
                    f"storage_location = '{loc_s}'",
                    f"location_id = COALESCE(location_id, "
                    f"  (SELECT location_id FROM stock_locations "
                    f"   WHERE location_code = '{loc_s}' LIMIT 1))",
                ]
            if stock_date:
                set_parts.append(
                    f"stock_date = COALESCE(stock_date, {stock_date_s})"
                )
            if mfg_date:
                set_parts.append(
                    f"mfg_date = COALESCE(mfg_date, {mfg_date_s})"
                )
            if remark1:
                set_parts.append(f"remark1 = COALESCE(remark1, {remark1_s})")
            if remark2:
                set_parts.append(f"remark2 = COALESCE(remark2, {remark2_s})")
            if package_box:
                set_parts.append(f"package_box = COALESCE(package_box, {pkg_box_s})")
            if bom_ref_val:
                set_parts.append(f"bom_ref = COALESCE(bom_ref, {bom_ref_s})")
            if unit_price_val is not None:
                set_parts.append(f"unit_price = COALESCE(unit_price, {unit_price_s})")

            sql_lines.append(
                f"UPDATE stock_items SET {', '.join(set_parts)} "
                f"WHERE d_id = '{d_id_s}';"
            )

            # ── 5. 歷史交易 INSERT ────────────────────────────────────────────
            running_stock = None
            for p in parsed_res:
                q_val   = p['qty']        if p['qty']        is not None else 0.0
                s_after = p['stock_after'] if p['stock_after'] is not None else 0.0

                if running_stock is not None:
                    q_before = running_stock
                else:
                    if   p['op'] == 'in':    q_before = s_after - q_val
                    elif p['op'] == 'out':   q_before = s_after + q_val
                    else:                    q_before = s_after

                if   p['op'] == 'out':                  signed_qty = -q_val
                elif p['op'] in ('count', 'adjust'):    signed_qty = s_after - q_before
                else:                                   signed_qty = q_val

                running_stock = s_after

                lf_s = p['loc_from'].replace("'", "''")
                lt_s = p['loc_to'].replace("'", "''")
                remark_s = p['raw'][:300].replace("'", "''")

                lf_id = (f"(SELECT location_id FROM stock_locations "
                         f"WHERE location_code = '{lf_s}' LIMIT 1)" if lf_s else 'NULL')
                lt_id = (f"(SELECT location_id FROM stock_locations "
                         f"WHERE location_code = '{lt_s}' LIMIT 1)" if lt_s else 'NULL')

                sql_lines.append(
                    f"INSERT INTO stock_transactions "
                    f"(stock_item_id, d_id, txn_type, txn_qty, qty_before, qty_after, "
                    f" location_from, location_to, location_from_id, location_to_id, "
                    f" txn_date, remark) "
                    f"VALUES ("
                    f"(SELECT stock_item_id FROM stock_items WHERE d_id = '{d_id_s}' LIMIT 1), "
                    f"'{d_id_s}', "
                    f"'{p['op']}', "
                    f"{int(signed_qty)}, "
                    f"{int(q_before)}, "
                    f"{int(s_after)}, "
                    f"'{lf_s}', '{lt_s}', "
                    f"{lf_id}, {lt_id}, "
                    f"'{p['date']}', '{remark_s}'"
                    f");"
                )
                total_ok += 1

            report['items'].append(item_rec)

    report['summary'] = {
        'total_items': len(report['items']),
        'total_transactions': total_ok,
        'parse_errors': total_error,
        'quantity_mismatches': 0,
        'sequence_warnings': 0
    }

    sql_lines.append("\nCOMMIT;")
    with open(OUTPUT_SQL,  'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_lines))
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print(f"Success! Processed {total_ok} transactions, {len(report['items'])} items.")

if __name__ == '__main__':
    main()
